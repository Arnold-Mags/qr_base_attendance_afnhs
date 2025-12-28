from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from .decorators import principal_required
from django.contrib import messages
from django.shortcuts import redirect
from .models import (
    AttendanceRecord,
    StudentProfile,
    Subject,
    DailyAttendance,
    SchoolSettings,
)
from .forms import SchoolSettingsForm
from django.db.models import Count, Q, Avg
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa


@principal_required
def principal_dashboard(request):
    """Principal dashboard with school-wide analytics and heatmaps"""
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday())
    month_start = today.replace(day=1)

    # School-wide statistics
    total_students = StudentProfile.objects.count()
    total_subjects = Subject.objects.count()

    # Today's attendance (Subject-based)
    today_records = AttendanceRecord.objects.filter(date=today)
    today_present = today_records.filter(Q(status="PRESENT") | Q(status="LATE")).count()
    today_absent = today_records.filter(status="ABSENT").count()
    today_total = today_records.count()

    # Today's General Attendance (Gate Logs)
    gate_records = DailyAttendance.objects.filter(date=today)
    gate_present = gate_records.filter(status="PRESENT").count()
    gate_late = gate_records.filter(status="LATE").count()
    gate_absent = total_students - (gate_present + gate_late)  # Approximation

    # For chart data (Gate Status)
    gate_stats = {"PRESENT": gate_present, "LATE": gate_late, "ABSENT": gate_absent}

    # Grade-level heatmap data (absence rates)
    grade_levels = [7, 8, 9, 10, 11, 12]
    heatmap_data = []

    for grade in grade_levels:
        students_in_grade = StudentProfile.objects.filter(grade_level=grade)
        total_in_grade = students_in_grade.count()

        if total_in_grade == 0:
            continue

        # Get this month's attendance for this grade
        month_records = AttendanceRecord.objects.filter(
            student__grade_level=grade, date__gte=month_start, date__lte=today
        )

        total_records = month_records.count()
        absent_records = month_records.filter(status="ABSENT").count()

        absence_rate = (
            (absent_records / total_records * 100) if total_records > 0 else 0
        )

        heatmap_data.append(
            {
                "grade_level": grade,
                "total_students": total_in_grade,
                "absence_rate": round(absence_rate, 1),
                "total_absences": absent_records,
            }
        )

    # Top 5 sections with highest absences
    sections_data = (
        AttendanceRecord.objects.filter(date__gte=month_start, status="ABSENT")
        .values("student__grade_level", "student__section")
        .annotate(absence_count=Count("id"))
        .order_by("-absence_count")[:5]
    )

    top_sections = []
    for section in sections_data:
        top_sections.append(
            {
                "grade_level": section["student__grade_level"],
                "section": section["student__section"],
                "absence_count": section["absence_count"],
            }
        )

    # Weekly trend data (for chart)
    weekly_trend = []
    for i in range(7):
        date = week_start + timedelta(days=i)
        if date > today:
            break

        day_records = AttendanceRecord.objects.filter(date=date)
        present = day_records.filter(Q(status="PRESENT") | Q(status="LATE")).count()
        absent = day_records.filter(status="ABSENT").count()

        weekly_trend.append(
            {
                "date": date.strftime("%a"),
                "present": present,
                "absent": absent,
                "total": present + absent,
            }
        )

    # Recent alerts (students marked absent today)
    recent_alerts = AttendanceRecord.objects.filter(
        date=today, status="ABSENT"
    ).select_related("student__user", "subject")[:10]

    context = {
        "total_students": total_students,
        "total_subjects": total_subjects,
        "today_present": today_present,
        "today_absent": today_absent,
        "today_total": today_total,
        "today_percentage": (
            round((today_present / today_total * 100), 1) if today_total > 0 else 0
        ),
        "heatmap_data": heatmap_data,
        "top_sections": top_sections,
        "weekly_trend": weekly_trend,
        "recent_alerts": recent_alerts,
        "current_month": today.strftime("%B %Y"),
        "gate_stats": gate_stats,
    }

    return render(request, "principal/dashboard.html", context)


@principal_required
def principal_settings_view(request):
    """View for principals to configure school-wide settings"""
    school_settings = SchoolSettings.get_settings()

    if request.method == "POST":
        form = SchoolSettingsForm(request.POST, request.FILES, instance=school_settings)
        if form.is_valid():
            form.save()
            messages.success(request, "School settings updated successfully.")
            return redirect("principal_settings")
    else:
        form = SchoolSettingsForm(instance=school_settings)

    context = {
        "form": form,
        "settings": school_settings,
    }
    return render(request, "principal/settings.html", context)


@principal_required
def attendance_logs_view(request):
    """View to display detailed attendance logs with filtering"""

    # Base Query
    logs = (
        DailyAttendance.objects.select_related("student", "student__user")
        .all()
        .order_by("-date", "-time_in_am")
    )

    # Filters
    search_query = request.GET.get("search", "")
    grade_filter = request.GET.get("grade", "")
    section_filter = request.GET.get("section", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    if search_query:
        logs = logs.filter(
            Q(student__user__first_name__icontains=search_query)
            | Q(student__user__last_name__icontains=search_query)
            | Q(student__student_id__icontains=search_query)
        )

    if grade_filter:
        logs = logs.filter(student__grade_level=grade_filter)

    if section_filter:
        logs = logs.filter(student__section__icontains=section_filter)

    if start_date:
        logs = logs.filter(date__gte=start_date)

    if end_date:
        logs = logs.filter(date__lte=end_date)

    if request.GET.get("export") == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Logs"

        # Headers
        headers = [
            "Date",
            "Student Name",
            "LRN",
            "Grade",
            "Section",
            "AM IN",
            "AM OUT",
            "PM IN",
            "PM OUT",
            "Status",
        ]
        ws.append(headers)

        # Style headers
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="10B981", end_color="10B981", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Data rows
        for log in logs:
            ws.append(
                [
                    log.date.strftime("%Y-%m-%d"),
                    log.student.user.get_full_name(),
                    log.student.student_id,
                    log.student.grade_level,
                    log.student.section,
                    log.time_in_am.strftime("%I:%M %p") if log.time_in_am else "--",
                    log.time_out_am.strftime("%I:%M %p") if log.time_out_am else "--",
                    log.time_in_pm.strftime("%I:%M %p") if log.time_in_pm else "--",
                    log.time_out_pm.strftime("%I:%M %p") if log.time_out_pm else "--",
                    log.status,
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="attendance_logs.xlsx"'
        wb.save(response)
        return response

    if request.GET.get("export") == "pdf":
        school_settings = SchoolSettings.get_settings()
        template_path = "principal/attendance_logs_pdf.html"
        context = {
            "logs": logs,
            "settings": school_settings,
            "grade_filter": grade_filter,
            "section_filter": section_filter,
            "start_date": start_date,
            "end_date": end_date,
            "generated_at": datetime.now(),
        }
        # Create a Django response object, and specify content_type as pdf
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="attendance_logs.pdf"'
        # find the template and render it.
        html = render_to_string(template_path, context)

        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response)
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse("We had some errors <pre>" + html + "</pre>")
        return response

    # Pagination
    from django.core.paginator import Paginator

    paginator = Paginator(logs, 50)  # 50 logs per page
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Context Data for dropdowns
    grade_levels = StudentProfile.GRADE_LEVEL_CHOICES
    # Get distinct sections? (Optional, might be slow if many students)
    # Ideally, we'd have a Section model, but for now we iterate unique sections from students
    sections = (
        StudentProfile.objects.values_list("section", flat=True)
        .distinct()
        .order_by("section")
    )

    context = {
        "logs": page_obj,
        "grade_levels": grade_levels,
        "sections": sections,
        # Maintain filter state
        "search_query": search_query,
        "grade_filter": grade_filter,
        "section_filter": section_filter,
        "start_date": start_date,
        "end_date": end_date,
    }

    return render(request, "principal/attendance_logs.html", context)
