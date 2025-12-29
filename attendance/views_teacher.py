from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from .decorators import teacher_required
from .models import (
    Subject,
    AttendanceRecord,
    StudentProfile,
    DailyAttendance,
    TeacherProfile,
    User,
)
from django.db.models import Count, Q
from datetime import datetime, timedelta, date as py_date
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import render_to_string
from xhtml2pdf import pisa
from .forms import StudentRegistrationForm, SubjectForm
from django.utils import timezone


@teacher_required
def teacher_reports_view(request):
    """View to generate attendance reports for subjects or advisory"""
    subjects = Subject.objects.filter(teacher=request.user)

    # Check advisory
    has_advisory = False
    advisory_grade = None
    advisory_section = None
    if hasattr(request.user, "teacher_profile"):
        profile = request.user.teacher_profile
        if profile.advisory_grade and profile.advisory_section:
            has_advisory = True
            advisory_grade = profile.advisory_grade
            advisory_section = profile.advisory_section

    # Filters
    report_type = request.GET.get("type", "daily")
    report_target = request.GET.get(
        "target",
        (
            f"subject_{subjects.first().id}"
            if subjects.exists()
            else "advisory" if has_advisory else ""
        ),
    )

    # Determine Date Range
    now = timezone.localtime()
    today = now.date()
    start_date = today
    end_date = today

    if report_type == "weekly":
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif report_type == "monthly":
        start_date = today.replace(day=1)
        end_date = today

    report_data = []
    is_advisory_report = False
    report_title = ""

    if report_target == "advisory" and has_advisory:
        is_advisory_report = True
        report_title = f"Advisory Class: Grade {advisory_grade} - {advisory_section}"
        students = (
            StudentProfile.objects.filter(
                grade_level=advisory_grade, section=advisory_section
            )
            .select_related("user")
            .order_by("user__last_name")
        )

        # Aggregate General (Gate) Attendance
        for student in students:
            records = DailyAttendance.objects.filter(
                student=student, date__gte=start_date, date__lte=end_date
            )
            present = records.filter(status="PRESENT").count()
            late = records.filter(status="LATE").count()
            absent = records.filter(status="ABSENT").count()
            total_days = (end_date - start_date).days + 1

            # Simple rate (present + late) / total
            rate = (
                round(((present + late) / total_days * 100), 1) if total_days > 0 else 0
            )

            report_data.append(
                {
                    "name": student.user.get_full_name(),
                    "lrn": student.student_id,
                    "present": present,
                    "late": late,
                    "absent": absent,
                    "total": total_days,
                    "rate": rate,
                }
            )

    elif report_target.startswith("subject_"):
        try:
            subject_id = report_target.split("_")[1]
            subject = subjects.get(id=subject_id)
            report_title = (
                f"Subject: {subject.name} (G{subject.grade_level}-{subject.section})"
            )
            students = (
                subject.students.all()
                .select_related("user")
                .order_by("user__last_name")
            )

            for student in students:
                records = AttendanceRecord.objects.filter(
                    student=student,
                    subject=subject,
                    date__gte=start_date,
                    date__lte=end_date,
                )
                present = records.filter(status="PRESENT").count()
                late = records.filter(status="LATE").count()
                absent = records.filter(status="ABSENT").count()
                total = records.count()

                rate = round(((present + late) / total * 100), 1) if total > 0 else 0

                report_data.append(
                    {
                        "name": student.user.get_full_name(),
                        "lrn": student.student_id,
                        "present": present,
                        "late": late,
                        "absent": absent,
                        "total": total,
                        "rate": rate,
                    }
                )
        except (Subject.DoesNotExist, IndexError, ValueError):
            pass

    # Handle Export
    if request.GET.get("export") == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Attendance Report"

        ws.append([report_title])
        ws.append([f"Report Type: {report_type.title()} ({start_date} to {end_date})"])
        ws.append([])  # spacer

        headers = [
            "Student Name",
            "LRN",
            "Present",
            "Late",
            "Absent",
            "Total",
            "Rate (%)",
        ]
        ws.append(headers)

        # Style headers
        for cell in ws[4]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="1F2937", end_color="1F2937", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        for row in report_data:
            ws.append(
                [
                    row["name"],
                    row["lrn"],
                    row["present"],
                    row["late"],
                    row["absent"],
                    row["total"],
                    row["rate"],
                ]
            )

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f'attachment; filename="Attendance_Report_{report_type}.xlsx"'
        )
        wb.save(response)
        return response

    if request.GET.get("export") == "pdf":
        template_path = "teacher/reports_pdf.html"
        context = {
            "report_title": report_title,
            "report_subtitle": f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
            "report_data": report_data,
            "is_advisory_report": is_advisory_report,
            "now": now,
        }
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = (
            f'attachment; filename="Attendance_Report_{report_type}.pdf"'
        )
        html = render_to_string(template_path, context)
        pisa_status = pisa.CreatePDF(html, dest=response)
        if pisa_status.err:
            return HttpResponse(f"Error generating PDF: {html}")
        return response

    context = {
        "subjects": subjects,
        "has_advisory": has_advisory,
        "advisory_grade": advisory_grade,
        "advisory_section": advisory_section,
        "report_type": report_type,
        "report_target": report_target,
        "report_title": report_title,
        "report_subtitle": f"{start_date.strftime('%B %d, %Y')} to {end_date.strftime('%B %d, %Y')}",
        "report_data": report_data,
        "is_advisory_report": is_advisory_report,
        "now": now,
    }
    return render(request, "teacher/reports.html", context)


def teacher_dashboard(request):
    """Teacher dashboard with assigned subjects and quick stats"""
    subjects = Subject.objects.filter(teacher=request.user).prefetch_related("students")

    subject_stats = []
    for subject in subjects:
        today = datetime.now().date()

        # Get today's attendance
        today_attendance = AttendanceRecord.objects.filter(
            subject=subject, date=today
        ).count()

        total_students = subject.students.count()

        # Get this week's stats
        week_start = today - timedelta(days=today.weekday())
        week_attendance = (
            AttendanceRecord.objects.filter(
                subject=subject, date__gte=week_start, date__lte=today
            )
            .values("status")
            .annotate(count=Count("id"))
        )

        week_stats = {"PRESENT": 0, "LATE": 0, "ABSENT": 0}
        for stat in week_attendance:
            week_stats[stat["status"]] = stat["count"]

        subject_stats.append(
            {
                "subject": subject,
                "total_students": total_students,
                "today_attendance": today_attendance,
                "week_stats": week_stats,
            }
        )

    context = {
        "subject_stats": subject_stats,
    }

    return render(request, "teacher/dashboard.html", context)


@teacher_required
def classroom_view(request, subject_id):
    """Classroom view with student photo grid and real-time attendance"""
    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)

    # Get all enrolled students
    students = subject.students.all().select_related("user")

    # Get today's attendance records
    today = datetime.now().date()
    attendance_records = AttendanceRecord.objects.filter(
        subject=subject, date=today
    ).select_related("student")

    # Create attendance map
    attendance_map = {record.student.id: record for record in attendance_records}

    # Prepare student data with attendance status
    student_data = []
    for student in students:
        attendance = attendance_map.get(student.id)
        student_data.append(
            {
                "student": student,
                "attendance": attendance,
                "status": attendance.status if attendance else "PENDING",
            }
        )

    # Statistics
    total_students = len(student_data)
    present_count = sum(1 for s in student_data if s["status"] == "PRESENT")
    late_count = sum(1 for s in student_data if s["status"] == "LATE")
    absent_count = sum(1 for s in student_data if s["status"] == "ABSENT")
    pending_count = sum(1 for s in student_data if s["status"] == "PENDING")

    context = {
        "subject": subject,
        "student_data": student_data,
        "total_students": total_students,
        "present_count": present_count,
        "late_count": late_count,
        "absent_count": absent_count,
        "pending_count": pending_count,
        "attendance_percentage": (
            round((present_count + late_count) / total_students * 100, 1)
            if total_students > 0
            else 0
        ),
    }

    return render(request, "teacher/classroom.html", context)


@teacher_required
def mark_manual_attendance(request, subject_id):
    """Manual attendance marking for teachers"""
    if request.method == "POST":
        subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)
        student_id = request.POST.get("student_id")
        status = request.POST.get("status")

        student = get_object_or_404(StudentProfile, id=student_id)

        # Check if student is enrolled
        if not subject.students.filter(id=student.id).exists():
            from django.contrib import messages

            messages.error(request, "Student not enrolled in this subject.")
            return redirect("classroom_view", subject_id=subject_id)

        # Create or update attendance record
        attendance, created = AttendanceRecord.objects.update_or_create(
            student=student,
            subject=subject,
            date=datetime.now().date(),
            defaults={"status": status, "scan_method": "MANUAL"},
        )

        from django.contrib import messages
        from .utils import trigger_parent_notification

        if status == "ABSENT":
            trigger_parent_notification(attendance)

        messages.success(
            request, f"Attendance marked for {student.user.get_full_name()}"
        )

        from django.shortcuts import redirect

        return redirect("classroom_view", subject_id=subject_id)

    return redirect("teacher_dashboard")


@teacher_required
def advisory_view(request):
    """View for teachers to monitor their advisory class daily attendance"""
    try:
        # Check if teacher profile exists
        if hasattr(request.user, "teacher_profile"):
            profile = request.user.teacher_profile
        else:
            # Auto-create if missing (optional, or handle as error)
            # For now, just handle as not assigned
            return render(request, "teacher/advisory.html", {"assigned": False})

        grade = profile.advisory_grade
        section = profile.advisory_section

        if not grade or not section:
            return render(request, "teacher/advisory.html", {"assigned": False})

        # Get students
        students = (
            StudentProfile.objects.filter(grade_level=grade, section=section)
            .select_related("user")
            .order_by("user__last_name")
        )

        today = datetime.now().date()
        daily_records = DailyAttendance.objects.filter(student__in=students, date=today)
        record_map = {r.student.id: r for r in daily_records}

        student_data = []
        present_count = 0
        late_count = 0

        for student in students:
            record = record_map.get(student.id)
            status = "PENDING"
            if record:
                status = record.status
                if status == "PRESENT":
                    present_count += 1
                if status == "LATE":
                    late_count += 1

            student_data.append(
                {"student": student, "record": record, "status": status}
            )

        context = {
            "assigned": True,
            "grade": grade,
            "section": section,
            "student_data": student_data,
            "today": today,
            "total_students": len(students),
            "present_count": present_count,
            "late_count": late_count,
            "absent_count": len(students)
            - present_count
            - late_count,  # Rough estimate
        }
        return render(request, "teacher/advisory.html", context)

    except Exception as e:
        # Log error?
        return render(
            request, "teacher/advisory.html", {"assigned": False, "error": str(e)}
        )


@teacher_required
def add_student_view(request):
    """View for teachers to add new students to their advisory class"""

    # Check assignment
    try:
        profile = request.user.teacher_profile
        if not profile.advisory_grade or not profile.advisory_section:
            from django.contrib import messages

            messages.error(request, "You do not have an advisory class assigned.")
            return redirect("teacher_dashboard")
    except TeacherProfile.DoesNotExist:
        from django.contrib import messages

        messages.error(request, "Teacher profile not found.")
        return redirect("teacher_dashboard")

    if request.method == "POST":
        form = StudentRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            # Create User
            username = form.cleaned_data["student_id"]  # Use LRN as username
            password = "password123"  # Default password (should be changed)
            first_name = form.cleaned_data["first_name"]
            last_name = form.cleaned_data["last_name"]

            # Check if user exists
            if User.objects.filter(username=username).exists():
                from django.contrib import messages

                messages.error(request, f"Student with LRN {username} already exists.")
            else:
                user = User.objects.create_user(
                    username=username,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                    role="STUDENT",
                )

                # Create Profile
                student = form.save(commit=False)
                student.user = user
                student.grade_level = profile.advisory_grade
                student.section = profile.advisory_section
                student.save()

                from django.contrib import messages

                messages.success(
                    request, f"Student {first_name} {last_name} added to your advisory."
                )
                return redirect("advisory_view")
    else:
        form = StudentRegistrationForm()

    context = {
        "form": form,
        "grade": profile.advisory_grade,
        "section": profile.advisory_section,
    }
    return render(request, "teacher/add_student.html", context)


@teacher_required
def add_subject_view(request):
    """View for teachers to add their own subjects"""
    if request.method == "POST":
        form = SubjectForm(request.POST)
        if form.is_valid():
            subject = form.save(commit=False)
            subject.teacher = request.user
            subject.save()
            from django.contrib import messages

            messages.success(request, f"Subject {subject.name} created successfully.")
            return redirect("teacher_dashboard")
    else:
        form = SubjectForm()

    return render(request, "teacher/add_subject.html", {"form": form})


@teacher_required
def edit_subject_view(request, subject_id):
    """View to edit an existing subject"""
    subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)

    if request.method == "POST":
        form = SubjectForm(request.POST, instance=subject)
        if form.is_valid():
            form.save()
            from django.contrib import messages

            messages.success(request, f"Subject {subject.name} updated successfully.")
            return redirect("teacher_dashboard")
    else:
        form = SubjectForm(instance=subject)

    return render(
        request,
        "teacher/add_subject.html",
        {"form": form, "is_edit": True, "subject": subject},
    )


@teacher_required
def delete_subject_view(request, subject_id):
    """View to delete a subject"""
    if request.method == "POST":
        subject = get_object_or_404(Subject, id=subject_id, teacher=request.user)
        subject_name = subject.name
        subject.delete()
        from django.contrib import messages

        messages.success(request, f"Subject {subject_name} deleted successfully.")
        return redirect("teacher_dashboard")

    return redirect("teacher_dashboard")
